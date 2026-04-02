"""
股票搜索框UI重叠Bug修复 - 测试用例生成脚本
项目: StockAnalysis
模块: 股票分析 > 搜索组件
日期: 2026-02-28
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_testcases():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "搜索组件UI测试"

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = ["用例编号", "所属模块", "用例名称", "优先级", "前置条件", "步骤描述", "预期结果"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 测试用例数据
    testcases = [
        # ==================== P0: 核心正向场景 ====================
        {
            "id": "TC_SEARCH_001",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索框无UI重叠，只显示单一边框层",
            "priority": "P0",
            "precondition": (
                "**环境要求**: 前端开发服务器(localhost:3000)正常运行，后端Mock服务(localhost:8000)正常。\n"
                "**开始路径**: 浏览器访问 http://localhost:3000/analysis\n"
                "**测试数据**: 无需额外准备"
            ),
            "steps": (
                "1. 打开股票分析页面\n"
                "2. 检验步骤1: 观察搜索框外观，检查是否存在多层嵌套边框\n"
                "3. 使用浏览器开发者工具(F12)检查搜索框DOM结构\n"
                "4. 检验步骤2: 逐层检查以下元素的border属性:\n"
                "   - .stock-search-container\n"
                "   - .ant-select-selector\n"
                "   - .ant-input-affix-wrapper\n"
                "   - .ant-input"
            ),
            "expected": (
                "检验步骤1: 搜索框视觉上只有一层边框，无多重嵌套矩形，无UI重叠现象。\n"
                "检验步骤2: 仅 .ant-input-affix-wrapper 有可见边框(1px solid)，"
                ".ant-select-selector 的 border 为 none，"
                ".ant-input 的 border 为 none。不存在多层边框叠加。"
            ),
        },
        {
            "id": "TC_SEARCH_002",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索框完整搜索流程正常工作",
            "priority": "P0",
            "precondition": (
                "**环境要求**: 前端开发服务器和Mock API服务正常运行。\n"
                "**开始路径**: 浏览器访问 http://localhost:3000/analysis\n"
                "**测试数据**: Mock服务包含股票数据(如'平安银行 000001.SZ')"
            ),
            "steps": (
                "1. 在搜索框中输入关键字'平安'\n"
                "2. 检验步骤1: 观察搜索框状态和下拉菜单\n"
                "3. 从下拉列表中点击选择'平安银行'\n"
                "4. 检验步骤2: 观察页面加载和显示结果"
            ),
            "expected": (
                "检验步骤1: 搜索框显示输入文字，下拉菜单出现匹配结果，"
                "下拉菜单无UI重叠，选项文字清晰可读。\n"
                "检验步骤2: 页面加载股票基本信息和K线图表，数据正确显示。"
            ),
        },

        # ==================== P1: 单功能点验证 ====================
        {
            "id": "TC_SEARCH_003",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索框Placeholder文字只显示一次无重叠",
            "priority": "P1",
            "precondition": (
                "**环境要求**: 前端服务正常。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 打开股票分析页面\n"
                "2. 确保搜索框处于空状态(未输入任何文字)\n"
                "3. 检验步骤1: 观察搜索框placeholder文字"
            ),
            "expected": (
                "检验步骤1: placeholder文字'请输入股票代码或名称进行搜索'只显示一次，"
                "无重影/重叠/模糊现象，颜色为浅灰色(--color-text-tertiary)。"
            ),
        },
        {
            "id": "TC_SEARCH_004",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索框聚焦状态边框变化正常",
            "priority": "P1",
            "precondition": (
                "**环境要求**: 前端服务正常。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 点击搜索框获取焦点\n"
                "2. 检验步骤1: 观察搜索框的边框和阴影变化\n"
                "3. 点击搜索框外部区域失去焦点\n"
                "4. 检验步骤2: 观察搜索框恢复状态"
            ),
            "expected": (
                "检验步骤1: 搜索框边框变为主题色(#1890ff)，"
                "出现蓝色发光阴影(box-shadow)，只有一层边框高亮，无多层边框同时高亮。\n"
                "检验步骤2: 搜索框恢复默认状态，边框恢复为 var(--glass-border) 颜色，"
                "阴影消失，无残留的多层边框。"
            ),
        },
        {
            "id": "TC_SEARCH_005",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索框悬停状态只有单层边框变化",
            "priority": "P1",
            "precondition": (
                "**环境要求**: 前端服务正常。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 将鼠标移入搜索框区域\n"
                "2. 检验步骤1: 观察搜索框悬停效果\n"
                "3. 将鼠标移出搜索框\n"
                "4. 检验步骤2: 观察搜索框恢复状态"
            ),
            "expected": (
                "检验步骤1: 搜索框背景微微变亮(rgba(255,255,255,0.08))，"
                "边框颜色变为主题色，只有一层边框变化，无内部嵌套边框同时响应hover。\n"
                "检验步骤2: 搜索框恢复默认外观，过渡平滑无闪烁。"
            ),
        },
        {
            "id": "TC_SEARCH_006",
            "module": "股票分析 > 搜索组件",
            "name": "验证下拉菜单样式与深色主题协调",
            "priority": "P1",
            "precondition": (
                "**环境要求**: 前端和Mock服务正常。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: Mock服务包含搜索数据"
            ),
            "steps": (
                "1. 在搜索框输入关键字触发下拉菜单\n"
                "2. 检验步骤1: 观察下拉菜单外观\n"
                "3. 将鼠标悬停在某个选项上\n"
                "4. 检验步骤2: 观察选项的悬停效果"
            ),
            "expected": (
                "检验步骤1: 下拉菜单背景为深色(--color-bg-elevated)，"
                "有模糊玻璃效果(backdrop-filter)，边框为 var(--glass-border)，"
                "文字颜色清晰可读。\n"
                "检验步骤2: 悬停选项背景高亮(蓝色半透明)，"
                "有向右偏移动画(translateX(4px))，过渡平滑。"
            ),
        },
        {
            "id": "TC_SEARCH_007",
            "module": "股票分析 > 搜索组件",
            "name": "验证清除按钮功能和样式正常",
            "priority": "P1",
            "precondition": (
                "**环境要求**: 前端服务正常。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 在搜索框输入任意文字\n"
                "2. 检验步骤1: 观察清除按钮是否出现\n"
                "3. 点击清除按钮\n"
                "4. 检验步骤2: 观察搜索框状态"
            ),
            "expected": (
                "检验步骤1: 输入框右侧出现清除按钮(X图标)，颜色为 var(--color-text-secondary)，"
                "鼠标悬停时变为主题色。\n"
                "检验步骤2: 输入框文字被清除，恢复显示placeholder，"
                "下拉菜单关闭，搜索框无多余边框残留。"
            ),
        },
        {
            "id": "TC_SEARCH_008",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索加载状态和空状态显示正确",
            "priority": "P1",
            "precondition": (
                "**环境要求**: 前端和Mock服务正常。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 在搜索框快速输入关键字\n"
                "2. 检验步骤1: 观察加载状态\n"
                "3. 输入不存在的股票名称(如'ZZZZZ')\n"
                "4. 等待搜索结果返回\n"
                "5. 检验步骤2: 观察空状态显示"
            ),
            "expected": (
                "检验步骤1: 下拉区域显示加载提示文字，无样式错乱。\n"
                "检验步骤2: 下拉区域显示空状态提示，文字居中，颜色为 var(--color-text-tertiary)。"
            ),
        },
        {
            "id": "TC_SEARCH_009",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索框未被Card组件包裹产生额外边框",
            "priority": "P1",
            "precondition": (
                "**环境要求**: 前端服务正常。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 打开浏览器开发者工具(F12)\n"
                "2. 定位搜索框组件的DOM结构\n"
                "3. 检验步骤1: 检查搜索框的父级元素"
            ),
            "expected": (
                "检验步骤1: 搜索框的直接父级为普通div元素(非ant-card)，"
                "无.glass-card类名，无.ant-card类名，"
                "父div仅有 style='margin-bottom: 24px' 属性。"
            ),
        },

        # ==================== P2: CSS属性精确验证 ====================
        {
            "id": "TC_SEARCH_010",
            "module": "股票分析 > 搜索组件",
            "name": "验证.ant-select-selector的border和background被正确清除",
            "priority": "P2",
            "precondition": (
                "**环境要求**: 前端服务正常，Chrome浏览器。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 打开Chrome开发者工具(F12)\n"
                "2. 使用元素选择器定位搜索框中的 .ant-select-selector 元素\n"
                "3. 检验步骤1: 在Computed面板中检查以下CSS属性:\n"
                "   - border\n"
                "   - background\n"
                "   - box-shadow\n"
                "   - padding"
            ),
            "expected": (
                "检验步骤1:\n"
                "- border: none (或 0px none)\n"
                "- background: transparent (或 rgba(0,0,0,0))\n"
                "- box-shadow: none\n"
                "- padding: 0px\n"
                "确认该元素不产生任何可见边框或背景。"
            ),
        },
        {
            "id": "TC_SEARCH_011",
            "module": "股票分析 > 搜索组件",
            "name": "验证.ant-input的border在搜索框上下文中被正确清除",
            "priority": "P2",
            "precondition": (
                "**环境要求**: 前端服务正常，Chrome浏览器。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 打开Chrome开发者工具(F12)\n"
                "2. 使用元素选择器定位 .stock-search-container 内的 .ant-input 元素\n"
                "3. 检验步骤1: 在Computed面板中检查以下CSS属性:\n"
                "   - border\n"
                "   - background\n"
                "   - box-shadow"
            ),
            "expected": (
                "检验步骤1:\n"
                "- border: none (或 0px none)\n"
                "- background: transparent (或 rgba(0,0,0,0))\n"
                "- box-shadow: none\n"
                "确认内部input元素不产生额外可见边框。"
            ),
        },
        {
            "id": "TC_SEARCH_012",
            "module": "股票分析 > 搜索组件",
            "name": "验证.ant-input-affix-wrapper作为唯一可见边框层的样式正确",
            "priority": "P2",
            "precondition": (
                "**环境要求**: 前端服务正常，Chrome浏览器。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 打开Chrome开发者工具(F12)\n"
                "2. 定位 .stock-search-container 内的 .ant-input-affix-wrapper 元素\n"
                "3. 检验步骤1: 在Computed面板中检查以下CSS属性:\n"
                "   - border\n"
                "   - background\n"
                "   - border-radius\n"
                "   - padding\n"
                "   - box-shadow"
            ),
            "expected": (
                "检验步骤1:\n"
                "- border: 1px solid (颜色接近 rgba(255,255,255,0.1))\n"
                "- background: 接近 rgba(255,255,255,0.05)\n"
                "- border-radius: 8px\n"
                "- padding: 12px 16px\n"
                "- box-shadow: 存在微弱阴影\n"
                "确认这是唯一具有可见边框的搜索框层。"
            ),
        },
        {
            "id": "TC_SEARCH_013",
            "module": "股票分析 > 搜索组件",
            "name": "验证搜索框在768px以下响应式适配正常",
            "priority": "P2",
            "precondition": (
                "**环境要求**: 前端服务正常，Chrome浏览器。\n"
                "**开始路径**: http://localhost:3000/analysis\n"
                "**测试数据**: 无"
            ),
            "steps": (
                "1. 打开Chrome开发者工具(F12)\n"
                "2. 切换到设备模拟器模式\n"
                "3. 将视口宽度设置为 375px (iPhone)\n"
                "4. 检验步骤1: 观察搜索框外观\n"
                "5. 输入关键字触发下拉菜单\n"
                "6. 检验步骤2: 观察下拉菜单"
            ),
            "expected": (
                "检验步骤1: 搜索框宽度自适应屏幕宽度，padding缩小为 10px 14px，"
                "字体大小调整为14px，仍然只有一层边框无重叠。\n"
                "检验步骤2: 下拉菜单最大高度调整为240px，选项正常显示，不溢出屏幕。"
            ),
        },
    ]

    # 写入测试用例数据
    for row_idx, tc in enumerate(testcases, 2):
        ws.cell(row=row_idx, column=1, value=tc["id"]).alignment = cell_alignment
        ws.cell(row=row_idx, column=2, value=tc["module"]).alignment = cell_alignment
        ws.cell(row=row_idx, column=3, value=tc["name"]).alignment = cell_alignment
        ws.cell(row=row_idx, column=4, value=tc["priority"]).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row_idx, column=5, value=tc["precondition"]).alignment = cell_alignment
        ws.cell(row=row_idx, column=6, value=tc["steps"]).alignment = cell_alignment
        ws.cell(row=row_idx, column=7, value=tc["expected"]).alignment = cell_alignment

        # 添加边框
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border

    # 设置列宽
    column_widths = {
        'A': 18,  # 用例编号
        'B': 22,  # 所属模块
        'C': 45,  # 用例名称
        'D': 10,  # 优先级
        'E': 55,  # 前置条件
        'F': 60,  # 步骤描述
        'G': 60,  # 预期结果
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 保存文件
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "StockAnalysis_搜索框UI修复_测试用例.xlsx")
    wb.save(output_path)
    print(f"测试用例已生成: {output_path}")
    print(f"共 {len(testcases)} 条用例 (P0: 2条, P1: 7条, P2: 4条)")

if __name__ == "__main__":
    create_testcases()
