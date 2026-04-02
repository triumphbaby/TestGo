# -*- coding: utf-8 -*-
"""
CopySource AI学习继承 - 测试用例生成脚本
按照 docs/testcase-spec.md 规范，生成四类测试用例 xlsx 文件
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 通用样式配置（遵循 testcase-spec.md）
# ============================================================

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGNMENT = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADERS = ["用例编号", "所属模块", "用例名称", "优先级", "前置条件", "步骤描述", "预期结果"]
COL_WIDTHS = [16, 22, 40, 8, 55, 55, 55]

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def add_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    # 默认隐藏用例编号列（A列）
    ws.column_dimensions['A'].hidden = True
    return ws


def save_workbook(wb, filename):
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"  -> {filepath}")


# ============================================================
# 1. 功能测试用例
# ============================================================

def gen_functional():
    wb = create_workbook()

    rows = [
        # ---- P0：核心端到端流程 ----
        [
            'TC_COPY_001',
            '文件管理 > 拷贝AI学习继承',
            '验证复制已完成AI学习的文件并粘贴后，副本自动继承学习状态无需重新学习',
            'P0',
            '**环境要求**：NAS 文件管理服务、AI 学习服务正常运行，存储池状态正常。\n'
            '**开始路径**：已使用管理员账号登录，进入【文件管理】页面。\n'
            '**测试数据**：目录 /testdata/ 下存在文件 learned_doc.pdf（已完成 AI 学习，AI 学习状态为「已完成」）。',

            '1. 在文件列表中选中 learned_doc.pdf。\n'
            '2. 点击【复制】按钮（或右键菜单选择「复制」）。\n'
            '3. 导航到目标目录 /testdata/backup/。\n'
            '4. 点击【粘贴】按钮。\n'
            '5. 等待粘贴操作完成。\n'
            '6. 检验步骤1：检查副本文件的 AI 学习状态。\n'
            '7. 检验步骤2：通过命令行检查副本文件的 CopySource 扩展属性。\n'
            '8. 检验步骤3：对副本文件执行 AI 相关功能（如语义搜索），验证可直接命中。',

            '检验步骤1：/testdata/backup/learned_doc.pdf 的 AI 学习状态显示为「已完成」，无「学习中」或「待学习」标记。\n'
            '检验步骤2：执行 getfattr 或等效命令，副本文件的 AttrCopySource 属性值为 /testdata/learned_doc.pdf<分隔符><源文件ID>，格式正确。\n'
            '检验步骤3：语义搜索可命中副本文件内容，返回结果中包含 /testdata/backup/learned_doc.pdf。'
        ],
        [
            'TC_COPY_002',
            '文件管理 > 拷贝AI学习继承',
            '验证复制未进行AI学习的文件并粘贴后，副本仍需正常触发AI学习流程',
            'P0',
            '**环境要求**：NAS 文件管理服务、AI 学习服务正常运行。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：目录 /testdata/ 下存在文件 new_doc.pdf（未进行 AI 学习，状态为「待学习」或无状态）。',

            '1. 选中 new_doc.pdf，执行复制操作。\n'
            '2. 导航到 /testdata/backup/，执行粘贴。\n'
            '3. 等待粘贴完成。\n'
            '4. 检验步骤1：检查副本文件的 AI 学习状态。\n'
            '5. 检验步骤2：检查副本文件的 CopySource 扩展属性。\n'
            '6. 检验步骤3：触发 AI 学习任务，观察副本文件是否被纳入学习队列。',

            '检验步骤1：副本文件 AI 学习状态为「待学习」或无状态，不显示「已完成」。\n'
            '检验步骤2：副本文件的 AttrCopySource 属性值包含源文件路径 /testdata/new_doc.pdf 及其文件ID。\n'
            '检验步骤3：AI 学习队列中包含该副本文件，学习任务正常启动。'
        ],

        # ---- P1：单功能点验证 ----
        [
            'TC_COPY_003',
            '文件管理 > GetCopySource',
            '验证对无 CopySource 属性的原始文件调用 GetCopySource 返回空值',
            'P1',
            '**环境要求**：文件管理服务正常。\n'
            '**开始路径**：通过 SSH 或 API 接口直接调用。\n'
            '**测试数据**：文件 /testdata/original.txt 存在，且从未被设置过 CopySource 扩展属性。',

            '1. 对 /testdata/original.txt 调用 GetCopySource 接口。\n'
            '2. 检验步骤1：检查返回值。',

            '检验步骤1：返回 sourcePath 为空字符串，fileId 为 0，error 为 nil。接口调用无报错，符合 ENODATA 场景的正常处理。'
        ],
        [
            'TC_COPY_004',
            '文件管理 > SetCopySource',
            '验证 SetCopySource 正确写入扩展属性并可通过 GetCopySource 读回',
            'P1',
            '**环境要求**：文件管理服务正常，文件系统支持扩展属性。\n'
            '**开始路径**：通过 API 接口调用。\n'
            '**测试数据**：文件 /testdata/target.txt 存在。源文件路径 /testdata/source.txt，源文件 ID 为 12345。',

            '1. 调用 SetCopySource(ctx, /testdata/target.txt, /testdata/source.txt, 12345)。\n'
            '2. 调用 GetCopySource(ctx, /testdata/target.txt)。\n'
            '3. 检验步骤1：检查 SetCopySource 的返回值。\n'
            '4. 检验步骤2：检查 GetCopySource 的返回值。\n'
            '5. 检验步骤3：通过命令行 getfattr 直接验证扩展属性原始值。',

            '检验步骤1：SetCopySource 返回 nil（无错误）。\n'
            '检验步骤2：GetCopySource 返回 sourcePath=/testdata/source.txt，fileId=12345，error=nil。\n'
            '检验步骤3：扩展属性原始值为 /testdata/source.txt<AttrSep>12345。'
        ],
        [
            'TC_COPY_005',
            '文件管理 > 拷贝AI学习继承',
            '验证对已学习文件进行多次复制粘贴后，所有副本均继承AI学习状态',
            'P1',
            '**环境要求**：文件管理服务、AI 学习服务正常。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：文件 learned_img.jpg（已完成 AI 学习）。3 个目标目录 /dir_a/、/dir_b/、/dir_c/。',

            '1. 选中 learned_img.jpg，执行复制。\n'
            '2. 分别粘贴到 /dir_a/、/dir_b/、/dir_c/。\n'
            '3. 检验步骤1：检查 3 个副本的 AI 学习状态。\n'
            '4. 检验步骤2：检查 3 个副本的 CopySource 属性。',

            '检验步骤1：3 个副本的 AI 学习状态均为「已完成」。\n'
            '检验步骤2：3 个副本的 CopySource 属性均指向同一源文件路径和文件ID。'
        ],
        [
            'TC_COPY_006',
            '文件管理 > 拷贝AI学习继承',
            '验证链式复制场景（A复制为B，B再复制为C）的AI学习状态继承',
            'P1',
            '**环境要求**：文件管理服务、AI 学习服务正常。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：文件 A.pdf（已完成 AI 学习）。',

            '1. 复制 A.pdf，粘贴为 B.pdf。\n'
            '2. 复制 B.pdf，粘贴为 C.pdf。\n'
            '3. 检验步骤1：检查 B.pdf 的 CopySource 属性和 AI 学习状态。\n'
            '4. 检验步骤2：检查 C.pdf 的 CopySource 属性和 AI 学习状态。\n'
            '5. 检验步骤3：对 C.pdf 执行语义搜索验证。',

            '检验步骤1：B.pdf 的 CopySource 指向 A.pdf，AI 学习状态为「已完成」。\n'
            '检验步骤2：C.pdf 的 CopySource 指向 B.pdf（或追溯到原始 A.pdf），AI 学习状态为「已完成」。\n'
            '检验步骤3：语义搜索可命中 C.pdf 内容。'
        ],
        [
            'TC_COPY_007',
            '文件管理 > 拷贝AI学习继承',
            '验证剪切（移动）已学习文件后AI学习状态保持不变',
            'P1',
            '**环境要求**：文件管理服务、AI 学习服务正常。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：文件 /dir_src/learned.pdf（已完成 AI 学习）。目标目录 /dir_dst/。',

            '1. 选中 learned.pdf，执行剪切操作。\n'
            '2. 导航到 /dir_dst/，执行粘贴。\n'
            '3. 检验步骤1：检查 /dir_dst/learned.pdf 的 AI 学习状态。\n'
            '4. 检验步骤2：检查原路径 /dir_src/learned.pdf 是否已不存在。\n'
            '5. 检验步骤3：语义搜索验证文件可被命中且路径为新路径。',

            '检验步骤1：移动后文件的 AI 学习状态仍为「已完成」，未丢失。\n'
            '检验步骤2：/dir_src/learned.pdf 不存在（已移走），文件管理列表中无残留。\n'
            '检验步骤3：语义搜索结果中文件路径为 /dir_dst/learned.pdf。'
        ],
        [
            'TC_COPY_008',
            '文件管理 > 拷贝AI学习继承',
            '验证复制已学习文件到同目录并重命名后AI学习状态正确继承',
            'P1',
            '**环境要求**：文件管理服务、AI 学习服务正常。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：文件 /testdata/report.docx（已完成 AI 学习）。',

            '1. 选中 report.docx，执行复制。\n'
            '2. 在同目录 /testdata/ 下粘贴（系统自动重命名为 report_copy.docx 或类似名称）。\n'
            '3. 检验步骤1：检查副本的 AI 学习状态。\n'
            '4. 检验步骤2：检查副本的 CopySource 属性。',

            '检验步骤1：report_copy.docx 的 AI 学习状态为「已完成」。\n'
            '检验步骤2：CopySource 属性指向 /testdata/report.docx 及其文件ID。'
        ],
        [
            'TC_COPY_009',
            '文件管理 > GetCopySource',
            '验证 GetCopySource 对属性格式异常的文件返回明确错误',
            'P1',
            '**环境要求**：文件管理服务正常，文件系统支持扩展属性。\n'
            '**开始路径**：通过 API 接口或命令行操作。\n'
            '**测试数据**：文件 /testdata/bad_attr.txt，手动设置其 AttrCopySource 属性为非法格式（如 invalid_no_separator，缺少分隔符）。',

            '1. 通过 setfattr 手动将 bad_attr.txt 的 AttrCopySource 设置为 invalid_no_separator。\n'
            '2. 调用 GetCopySource(ctx, /testdata/bad_attr.txt)。\n'
            '3. 检验步骤1：检查返回值和错误信息。',

            '检验步骤1：返回 sourcePath 为空，fileId 为 0，error 不为 nil 且错误信息包含 invalid attr format。'
        ],
        [
            'TC_COPY_010',
            '文件管理 > GetCopySource',
            '验证 GetCopySource 对 fileId 非数字的属性值返回解析错误',
            'P1',
            '**环境要求**：文件管理服务正常。\n'
            '**开始路径**：通过 API 接口或命令行操作。\n'
            '**测试数据**：文件 /testdata/bad_id.txt，手动设置 AttrCopySource 为 /some/path<AttrSep>not_a_number。',

            '1. 手动设置 bad_id.txt 的 AttrCopySource 为 /some/path<AttrSep>not_a_number。\n'
            '2. 调用 GetCopySource(ctx, /testdata/bad_id.txt)。\n'
            '3. 检验步骤1：检查返回值和错误信息。',

            '检验步骤1：返回 error 不为 nil，错误信息包含 parse file id，sourcePath 为空，fileId 为 0。'
        ],

        # ---- P2：异常与容错 ----
        [
            'TC_COPY_011',
            '文件管理 > 拷贝AI学习继承',
            '验证源文件被删除后副本的AI学习状态和CopySource属性不受影响',
            'P2',
            '**环境要求**：文件管理服务、AI 学习服务正常。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：文件 /testdata/to_delete.pdf（已完成 AI 学习），已复制一份副本 /testdata/backup/to_delete.pdf。',

            '1. 确认副本 AI 学习状态为「已完成」且 CopySource 属性指向源文件。\n'
            '2. 删除源文件 /testdata/to_delete.pdf。\n'
            '3. 检验步骤1：检查副本的 AI 学习状态。\n'
            '4. 检验步骤2：检查副本的 CopySource 扩展属性是否仍存在。\n'
            '5. 检验步骤3：对副本执行语义搜索。',

            '检验步骤1：副本 AI 学习状态仍为「已完成」，不因源文件删除而丢失。\n'
            '检验步骤2：CopySource 属性仍存在，值未被清除（源路径仍记录在属性中）。\n'
            '检验步骤3：语义搜索仍可命中副本文件内容。'
        ],
        [
            'TC_COPY_012',
            '文件管理 > 拷贝AI学习继承',
            '验证复制操作中途设备重启后文件和属性的一致性',
            'P2',
            '**环境要求**：文件管理服务正常，至少有一个可用存储池。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：大文件 /testdata/large_learned.zip（500MB，已完成 AI 学习）。',

            '1. 发起复制 large_learned.zip 到 /testdata/backup/ 的粘贴操作。\n'
            '2. 在复制过程中（进度未达100%）执行设备重启。\n'
            '3. 等待设备重启完成并重新登录。\n'
            '4. 检验步骤1：检查目标目录中副本文件状态。\n'
            '5. 检验步骤2：检查系统日志。',

            '检验步骤1：副本文件不存在或状态标记为异常/不完整，不应显示为正常且 AI 学习状态为「已完成」。无脏数据残留。\n'
            '检验步骤2：系统日志中记录复制中断相关信息，无未捕获异常。'
        ],
    ]

    add_sheet(wb, '测试用例', rows)
    save_workbook(wb, 'CopySource_AI学习继承_功能测试用例.xlsx')


# ============================================================
# 2. 兼容性测试用例
# ============================================================

def gen_compatibility():
    wb = create_workbook()

    rows = [
        [
            'TC_COMPAT_001',
            '兼容性 > 文件类型',
            '验证复制已学习的不同类型文件（图片/文档/视频）后AI学习状态均正确继承',
            'P1',
            '**环境要求**：文件管理服务、AI 学习服务正常，Chrome 最新稳定版，局域网直连。\n'
            '**开始路径**：已登录管理员账号，进入【文件管理】页面。\n'
            '**测试数据**：准备以下已完成 AI 学习的文件：photo.jpg（3MB 图片）、document.pdf（2MB 文档）、video.mp4（50MB 视频）。',

            '1. 分别对 photo.jpg、document.pdf、video.mp4 执行复制 -> 粘贴到 /backup/ 目录。\n'
            '2. 检验步骤1：检查 photo.jpg 副本的 AI 学习状态和 CopySource 属性。\n'
            '3. 检验步骤2：检查 document.pdf 副本的 AI 学习状态和 CopySource 属性。\n'
            '4. 检验步骤3：检查 video.mp4 副本的 AI 学习状态和 CopySource 属性。',

            '检验步骤1：photo.jpg 副本 AI 学习状态为「已完成」，CopySource 指向原始路径。\n'
            '检验步骤2：document.pdf 副本 AI 学习状态为「已完成」，CopySource 指向原始路径。\n'
            '检验步骤3：video.mp4 副本 AI 学习状态为「已完成」，CopySource 指向原始路径。'
        ],
        [
            'TC_COMPAT_002',
            '兼容性 > 硬件平台',
            '验证 Ultra 机型上复制已学习文件的AI学习继承功能正常',
            'P1',
            '**环境要求**：Ultra 机型 NAS 设备，文件管理服务和 AI 学习服务正常，Chrome 最新稳定版，局域网直连，分辨率 1920x1080。\n'
            '**开始路径**：已登录管理员账号，进入【文件管理】页面。\n'
            '**测试数据**：已完成 AI 学习的文件 test_ultra.pdf。',

            '1. 复制 test_ultra.pdf 到 /backup/ 目录。\n'
            '2. 检验步骤1：检查副本 AI 学习状态。\n'
            '3. 检验步骤2：通过 SSH 检查扩展属性。\n'
            '4. 检验步骤3：语义搜索验证。',

            '检验步骤1：副本 AI 学习状态为「已完成」，与 x86 标准机型表现一致。\n'
            '检验步骤2：CopySource 扩展属性格式和值与标准表现一致。\n'
            '检验步骤3：语义搜索可命中副本，搜索性能与标准环境无明显差异。'
        ],
        [
            'TC_COMPAT_003',
            '兼容性 > 硬件平台',
            '验证 ARM 机型上复制已学习文件的AI学习继承功能正常',
            'P1',
            '**环境要求**：ARM 机型 NAS 设备，文件管理服务和 AI 学习服务正常，Chrome 最新稳定版，局域网直连。\n'
            '**开始路径**：已登录管理员账号，进入【文件管理】页面。\n'
            '**测试数据**：已完成 AI 学习的文件 test_arm.pdf。',

            '1. 复制 test_arm.pdf 到 /backup/ 目录。\n'
            '2. 检验步骤1：检查副本 AI 学习状态。\n'
            '3. 检验步骤2：通过 SSH 检查扩展属性值的字节序和编码。',

            '检验步骤1：副本 AI 学习状态为「已完成」，功能与 Ultra 机型一致。\n'
            '检验步骤2：CopySource 属性值格式正确，路径字符串和文件ID解析无误（ARM 字节序不影响属性读写）。'
        ],
        [
            'TC_COMPAT_004',
            '兼容性 > 浏览器',
            '验证 Firefox 浏览器下复制粘贴已学习文件后AI学习状态正确继承',
            'P1',
            '**环境要求**：Firefox 最新稳定版，分辨率 1920x1080，局域网直连。\n'
            '**开始路径**：使用 Firefox 打开 NAS 登录页并登录管理员账号，进入【文件管理】页面。\n'
            '**测试数据**：已完成 AI 学习的文件 test_firefox.pdf。',

            '1. 复制 test_firefox.pdf 到 /backup/ 目录。\n'
            '2. 检验步骤1：检查副本 AI 学习状态。\n'
            '3. 检验步骤2：检查页面 UI 表现（AI 状态图标/标签显示）。',

            '检验步骤1：副本 AI 学习状态为「已完成」，与 Chrome 下表现一致。\n'
            '检验步骤2：AI 学习状态图标/标签正常显示，无样式错位或缺失。'
        ],
        [
            'TC_COMPAT_005',
            '兼容性 > 网络环境',
            '验证通过公网穿透访问时复制已学习文件的AI学习继承功能正常',
            'P1',
            '**环境要求**：NAS 已配置公网穿透服务，通过外网域名可访问文件管理。Chrome 最新稳定版。\n'
            '**开始路径**：通过公网穿透域名登录管理员账号，进入【文件管理】页面。\n'
            '**测试数据**：已完成 AI 学习的文件 test_remote.pdf。',

            '1. 通过公网穿透地址复制 test_remote.pdf 到 /backup/。\n'
            '2. 检验步骤1：检查副本 AI 学习状态。\n'
            '3. 检验步骤2：检查复制操作耗时。',

            '检验步骤1：副本 AI 学习状态为「已完成」，功能与局域网访问一致。\n'
            '检验步骤2：复制操作正常完成，扩展属性写入未因网络延迟而失败。'
        ],
        [
            'TC_COMPAT_006',
            '兼容性 > 弱网',
            '验证弱网环境下复制已学习的大文件时AI学习状态继承和超时处理',
            'P2',
            '**环境要求**：使用 tc 工具模拟弱网（下行 1Mbps，上行 500kbps，延迟 500ms）。Chrome 最新稳定版。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：已完成 AI 学习的大文件 large_learned.zip（100MB）。',

            '1. 在弱网条件下复制 large_learned.zip 到 /backup/。\n'
            '2. 观察复制进度。\n'
            '3. 检验步骤1：复制完成后检查副本 AI 学习状态。\n'
            '4. 检验步骤2：若复制超时，检查错误提示和文件一致性。',

            '检验步骤1：若复制成功完成，副本 AI 学习状态为「已完成」，CopySource 属性正确。\n'
            '检验步骤2：若超时中断，前端显示明确的超时错误提示，目标目录无不完整的残留文件。'
        ],
    ]

    add_sheet(wb, '测试用例', rows)
    save_workbook(wb, 'CopySource_AI学习继承_兼容性测试用例.xlsx')


# ============================================================
# 3. 专项测试用例
# ============================================================

def gen_special():
    wb = create_workbook()

    storage_rows = [
        [
            'TC_STOR_001',
            '专项 > 存储异常 > 拔盘',
            '验证存储池降级状态下复制已学习文件的AI学习继承功能和数据完整性',
            'P2',
            '**环境要求**：已创建 RAID1 存储池 Pool_A（磁盘1+磁盘2），存储管理服务正常。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：Pool_A 中存在已完成 AI 学习的文件 learned_in_raid.pdf。物理拔出磁盘2使存储池降级。',

            '1. 物理拔出磁盘2，等待存储池状态变为「降级」。\n'
            '2. 在降级状态下复制 learned_in_raid.pdf 到 Pool_A 内另一目录。\n'
            '3. 检验步骤1：检查复制操作是否成功完成。\n'
            '4. 检验步骤2：检查副本的 CopySource 扩展属性。\n'
            '5. 检验步骤3：检查副本的 AI 学习状态。\n'
            '6. 检验步骤4：检查副本文件内容完整性（MD5校验）。',

            '检验步骤1：复制操作成功完成，无错误提示。\n'
            '检验步骤2：CopySource 属性正确写入，格式和值无异常。\n'
            '检验步骤3：AI 学习状态为「已完成」。\n'
            '检验步骤4：副本 MD5 与源文件一致，文件内容无损坏。'
        ],
        [
            'TC_STOR_002',
            '专项 > 存储异常 > 存储池损毁',
            '验证存储池损毁后文件CopySource扩展属性不可用时的错误处理',
            'P2',
            '**环境要求**：已创建 RAID0 存储池 Pool_B（2块盘，无冗余）。\n'
            '**开始路径**：通过 API 接口调用 GetCopySource。\n'
            '**测试数据**：Pool_B 中存在有 CopySource 属性的文件。拔出一块盘使池损毁。',

            '1. 拔出磁盘使 Pool_B 状态变为「损毁」。\n'
            '2. 调用 GetCopySource 尝试读取损毁池中文件的属性。\n'
            '3. 检验步骤1：检查接口返回。\n'
            '4. 检验步骤2：检查系统日志。',

            '检验步骤1：接口返回明确错误（非 panic），错误信息指示存储不可用，不返回脏数据。\n'
            '检验步骤2：系统日志中记录读取失败信息，级别为 ERROR，无未捕获异常。'
        ],
    ]

    security_rows = [
        [
            'TC_SEC_001',
            '专项 > 安全 > 属性篡改',
            '验证手动篡改CopySource扩展属性为恶意路径不会导致AI学习状态异常继承',
            'P2',
            '**环境要求**：文件管理服务、AI 学习服务正常。具备 SSH 访问权限。\n'
            '**开始路径**：通过 SSH 操作扩展属性，再通过 Web UI 验证。\n'
            '**测试数据**：文件 /testdata/tampered.txt（未进行 AI 学习）。一个不存在的伪造源路径 /fake/path/learned.pdf。',

            '1. 通过 SSH 手动设置 tampered.txt 的 AttrCopySource 为 /fake/path/learned.pdf<AttrSep>99999。\n'
            '2. 在 Web UI 中查看 tampered.txt 的 AI 学习状态。\n'
            '3. 检验步骤1：检查 AI 学习状态是否为「已完成」。\n'
            '4. 检验步骤2：检查系统是否对不存在的源路径进行了校验。',

            '检验步骤1：tampered.txt 的 AI 学习状态不为「已完成」（系统应校验源文件实际存在且确实已学习，而非仅信任属性值）。\n'
            '检验步骤2：系统日志中无因伪造路径导致的未处理异常；若有校验逻辑，记录 WARN 级别日志。'
        ],
        [
            'TC_SEC_002',
            '专项 > 安全 > 注入攻击',
            '验证CopySource属性中包含特殊字符（路径遍历/SQL注入）时系统安全处理',
            'P2',
            '**环境要求**：文件管理服务正常。具备 SSH 访问权限。\n'
            '**开始路径**：通过 SSH 和 API 接口操作。\n'
            '**测试数据**：文件 /testdata/inject_test.txt。恶意路径字符串：/../../../etc/passwd<AttrSep>1 和 \' OR 1=1 --<AttrSep>1。',

            '1. 通过 setfattr 将 inject_test.txt 的 AttrCopySource 设置为 /../../../etc/passwd<AttrSep>1。\n'
            '2. 调用 GetCopySource 并观察返回。\n'
            '3. 将属性值改为 \' OR 1=1 --<AttrSep>1。\n'
            '4. 调用 GetCopySource 并观察返回。\n'
            '5. 检验步骤1：检查路径遍历场景的返回值和系统行为。\n'
            '6. 检验步骤2：检查 SQL 注入场景的返回值和数据库状态。',

            '检验步骤1：GetCopySource 正常返回路径字符串（不执行路径遍历），系统后续使用该路径时应做路径规范化和权限校验，不返回 /etc/passwd 内容。\n'
            '检验步骤2：GetCopySource 正常返回字符串（不执行 SQL），数据库无异常查询日志，无数据泄露。'
        ],
        [
            'TC_SEC_003',
            '专项 > 安全 > 鉴权',
            '验证普通用户无法通过API篡改其他用户文件的CopySource属性',
            'P2',
            '**环境要求**：文件管理服务正常，系统中存在管理员账号和普通用户账号。\n'
            '**开始路径**：使用普通用户账号登录并通过 API 发起请求。\n'
            '**测试数据**：管理员的文件 /admin_files/secret.pdf（已完成 AI 学习）。普通用户无权限访问 /admin_files/ 目录。',

            '1. 使用普通用户的 Token 调用 SetCopySource API，目标路径为 /admin_files/secret.pdf。\n'
            '2. 检验步骤1：检查 API 返回状态码和错误信息。\n'
            '3. 检验步骤2：通过管理员账号检查 secret.pdf 的 CopySource 属性是否被篡改。',

            '检验步骤1：API 返回 HTTP 403 Forbidden 或 401 Unauthorized，响应体不包含任何文件内容或属性信息。\n'
            '检验步骤2：secret.pdf 的 CopySource 属性未被修改，文件元数据完整。'
        ],
    ]

    performance_rows = [
        [
            'TC_PERF_001',
            '专项 > 性能 > 批量复制',
            '验证批量复制100个已学习文件时CopySource属性写入的性能和正确性',
            'P2',
            '**环境要求**：文件管理服务正常，存储池剩余空间充足。\n'
            '**开始路径**：已登录，进入【文件管理】页面。\n'
            '**测试数据**：目录 /testdata/batch/ 下100个已完成 AI 学习的文件（混合类型，单文件大小 1-10MB）。',

            '1. 全选 /testdata/batch/ 下100个文件，执行复制。\n'
            '2. 粘贴到 /testdata/batch_backup/。\n'
            '3. 记录从开始粘贴到全部完成的总耗时。\n'
            '4. 检验步骤1：检查批量复制总耗时。\n'
            '5. 检验步骤2：逐一检查100个副本的 CopySource 属性。\n'
            '6. 检验步骤3：抽查10个副本的 AI 学习状态。',

            '检验步骤1：批量复制含属性写入总耗时与不含 AI 学习的同等操作相比，额外开销 < 10%。\n'
            '检验步骤2：100个副本的 CopySource 属性均正确，无遗漏或格式错误。\n'
            '检验步骤3：抽查的10个副本 AI 学习状态均为「已完成」。'
        ],
        [
            'TC_PERF_002',
            '专项 > 性能 > 并发读取',
            '验证多用户并发调用GetCopySource的响应时间和系统稳定性',
            'P2',
            '**环境要求**：文件管理服务正常。可使用压测工具（如 wrk/ab）。\n'
            '**开始路径**：通过压测工具并发调用 API。\n'
            '**测试数据**：文件 /testdata/concurrent_test.pdf 已设置 CopySource 属性。模拟 50 并发用户。',

            '1. 使用压测工具对 GetCopySource API 发起 50 并发请求，持续 30 秒。\n'
            '2. 检验步骤1：检查 API 平均响应时间。\n'
            '3. 检验步骤2：检查请求成功率。\n'
            '4. 检验步骤3：检查系统资源占用。',

            '检验步骤1：平均响应时间 < 100ms，P99 响应时间 < 500ms。\n'
            '检验步骤2：请求成功率 100%，无 5xx 错误。\n'
            '检验步骤3：CPU 占用无异常飙升（< 80%），内存无持续增长趋势。'
        ],
    ]

    add_sheet(wb, '存储异常', storage_rows)
    add_sheet(wb, '安全测试', security_rows)
    add_sheet(wb, '性能压力', performance_rows)
    save_workbook(wb, 'CopySource_AI学习继承_专项测试用例.xlsx')


# ============================================================
# 4. 升级测试用例
# ============================================================

def gen_upgrade():
    wb = create_workbook()

    rows = [
        [
            'TC_UPG_001',
            '升级测试 > OTA',
            '验证OTA升级后已有文件的CopySource扩展属性和AI学习状态完整保留',
            'P1',
            '**环境要求**：当前系统版本 vN（支持 CopySource 特性），目标版本 vN+1，OTA 服务正常。\n'
            '**开始路径**：已登录管理员账号，进入【系统设置】->【系统升级】页面。\n'
            '**测试数据**：系统中存在：(1)已完成 AI 学习的原始文件 original.pdf；(2)通过复制生成的副本 copy.pdf（CopySource 指向 original.pdf，AI 学习状态为「已完成」）。记录升级前 copy.pdf 的 CopySource 属性值。',

            '1. 记录升级前 copy.pdf 的 CopySource 属性值和 AI 学习状态。\n'
            '2. 执行 OTA 升级到 vN+1，等待重启完成。\n'
            '3. 重新登录系统。\n'
            '4. 检验步骤1：检查系统版本号。\n'
            '5. 检验步骤2：通过 SSH 检查 copy.pdf 的 CopySource 扩展属性。\n'
            '6. 检验步骤3：在 Web UI 检查 copy.pdf 的 AI 学习状态。\n'
            '7. 检验步骤4：对 copy.pdf 执行语义搜索。',

            '检验步骤1：系统版本为 vN+1。\n'
            '检验步骤2：CopySource 属性值与升级前记录完全一致，格式无损。\n'
            '检验步骤3：AI 学习状态仍为「已完成」。\n'
            '检验步骤4：语义搜索可正常命中 copy.pdf。'
        ],
        [
            'TC_UPG_002',
            '升级测试 > OTA',
            '验证从不支持CopySource的旧版本升级后，旧文件正常工作且新复制可启用该特性',
            'P1',
            '**环境要求**：当前系统版本 vOLD（不支持 CopySource 特性），目标版本 vNEW（支持该特性），OTA 服务正常。\n'
            '**开始路径**：已登录管理员账号。\n'
            '**测试数据**：旧版本中已存在已完成 AI 学习的文件 old_learned.pdf（无 CopySource 扩展属性）。',

            '1. 从 vOLD 升级到 vNEW。\n'
            '2. 检验步骤1：调用 GetCopySource 查询 old_learned.pdf。\n'
            '3. 在新版本中复制 old_learned.pdf 到 /backup/。\n'
            '4. 检验步骤2：检查副本的 CopySource 属性。\n'
            '5. 检验步骤3：检查副本的 AI 学习状态。',

            '检验步骤1：GetCopySource 返回空值（sourcePath 为空, fileId=0, error=nil），符合 ENODATA 场景，无报错。\n'
            '检验步骤2：副本的 CopySource 属性正确设置，指向 old_learned.pdf 路径和文件ID。\n'
            '检验步骤3：副本的 AI 学习状态为「已完成」，新特性在旧文件上正常生效。'
        ],
        [
            'TC_UPG_003',
            '升级测试 > 数据库变更',
            '验证升级后AI学习状态表与CopySource属性的数据一致性',
            'P1',
            '**环境要求**：当前版本 vN，目标版本 vN+1（AI 学习表可能有结构变更）。\n'
            '**开始路径**：通过 SSH 连接数据库和文件系统。\n'
            '**测试数据**：升级前系统中存在 10 个有 CopySource 属性的副本文件，其中 8 个源文件仍存在、2 个源文件已删除。',

            '1. 升级前记录 10 个副本文件的 CopySource 属性值和对应 AI 学习数据库记录。\n'
            '2. 执行 OTA 升级。\n'
            '3. 检验步骤1：查询数据库 AI 学习状态表，检查 10 个副本的记录。\n'
            '4. 检验步骤2：检查 10 个副本文件的 CopySource 扩展属性。\n'
            '5. 检验步骤3：检查源文件已删除的 2 个副本的 AI 学习状态。',

            '检验步骤1：数据库中 10 个副本的 AI 学习记录完整保留，字段值与升级前一致。\n'
            '检验步骤2：10 个副本的 CopySource 扩展属性值与升级前记录完全一致。\n'
            '检验步骤3：源文件已删除的 2 个副本 AI 学习状态仍为「已完成」（状态独立于源文件存在性）。'
        ],
        [
            'TC_UPG_004',
            '升级测试 > 跨版本升级',
            '验证跨版本升级（vN-2到vN）后CopySource属性和AI学习状态兼容性',
            'P1',
            '**环境要求**：当前版本 vN-2，目标版本 vN（跳过 vN-1），OTA 服务正常。\n'
            '**开始路径**：已登录管理员账号，进入【系统升级】页面。\n'
            '**测试数据**：vN-2 版本中存在已完成 AI 学习的文件及其副本（若 vN-2 支持 CopySource），或仅存在已学习文件（若 vN-2 不支持）。',

            '1. 执行跨版本升级从 vN-2 直接到 vN。\n'
            '2. 检验步骤1：系统升级是否成功完成。\n'
            '3. 检验步骤2：检查已有文件的 AI 学习状态。\n'
            '4. 检验步骤3：在 vN 上执行复制已学习文件操作，验证 CopySource 特性。\n'
            '5. 检验步骤4：检查数据库迁移脚本执行日志。',

            '检验步骤1：升级成功完成，无中断或报错。\n'
            '检验步骤2：已有文件的 AI 学习状态保持不变。\n'
            '检验步骤3：新复制的副本正确设置 CopySource 属性，AI 学习状态为「已完成」。\n'
            '检验步骤4：数据库迁移日志无 ERROR，所有迁移脚本按序执行成功。'
        ],
        [
            'TC_UPG_005',
            '升级测试 > 刷机初始化',
            '验证刷机初始化后CopySource特性可正常使用',
            'P1',
            '**环境要求**：已准备 vN 刷机固件，设备可通过刷机工具连接。\n'
            '**开始路径**：使用刷机工具将固件写入设备。\n'
            '**测试数据**：无（全新初始化状态）。',

            '1. 刷入 vN 固件，完成初始化引导。\n'
            '2. 上传一个文件 test_init.pdf 并触发 AI 学习，等待学习完成。\n'
            '3. 复制 test_init.pdf 到 /backup/。\n'
            '4. 检验步骤1：检查副本的 CopySource 属性。\n'
            '5. 检验步骤2：检查副本的 AI 学习状态。\n'
            '6. 检验步骤3：语义搜索验证。',

            '检验步骤1：CopySource 属性正确设置，指向源文件路径和文件ID。\n'
            '检验步骤2：AI 学习状态为「已完成」。\n'
            '检验步骤3：语义搜索可命中副本文件内容。'
        ],
    ]

    add_sheet(wb, '测试用例', rows)
    save_workbook(wb, 'CopySource_AI学习继承_升级测试用例.xlsx')


# ============================================================
# 主函数
# ============================================================

if __name__ == '__main__':
    print('正在生成测试用例文件...\n')
    gen_functional()
    gen_compatibility()
    gen_special()
    gen_upgrade()
    print('\n全部生成完成！')
